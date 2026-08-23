"""Mirrors the SQLite data into a single Excel workbook (instance/erp_data.xlsx).

Call sync_to_excel() right after any db.session.commit() that adds, edits, or
deletes a row. It re-reads every table from the DB and rewrites the workbook,
so the two are always consistent with each other — there's no incremental
diffing to get wrong. The workbook is a read-only mirror for the shop owner
to browse/filter in Excel; the database remains the single source of truth.

Writing is best-effort: if erp_data.xlsx is currently open in Excel (locked
on Windows) or anything else goes wrong, the DB write already succeeded, so
we just flash a warning instead of failing the request.
"""

import os
import shutil
from datetime import datetime

from flask import current_app, flash
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from models import (
    Product, Supplier, Customer, CustomerBrandDiscount, Mechanic, MechanicBrandDiscount,
    Purchase, PurchaseItem, Sale, SaleItem, StockMovement, ShopSettings, Payment, Brand,
    SaleReturn, SaleReturnItem,
)

EXCEL_FILENAME = "erp_data.xlsx"

# sheet_name -> (model, [column headers])
# Columns that aren't plain model attributes (e.g. "supplier_name") are
# resolved by _cell_value below.
SHEETS = {
    "Products": (Product, [
        "id", "part_no", "product_name", "brand_name", "category", "vehicle_name", "unit", "hsn_code",
        "gst_rate", "mrp", "actual_discount_pct", "actual_discounted_price",
        "margin_per_unit", "current_stock", "reorder_level",
    ]),
    "Suppliers": (Supplier, ["id", "name", "brand_name", "phone", "address", "gstin"]),
    "Brands": (Brand, ["id", "name"]),
    "Customers": (Customer, ["id", "name", "phone", "address", "vehicle_model"]),
    "Mechanics": (Mechanic, ["id", "name", "phone", "garage_name"]),
    "Purchases": (Purchase, [
        "id", "date", "invoice_no", "supplier_id", "supplier_name", "total",
    ]),
    "PurchaseItems": (PurchaseItem, [
        "id", "purchase_id", "product_id", "product_name", "qty", "purchase_price",
        "mrp_at_purchase", "remaining_qty", "stock_number", "total",
    ]),
    "Sales": (Sale, [
        "id", "date", "invoice_no", "customer_id", "customer_name", "mechanic_id",
        "mechanic_name", "payment_mode", "amount_paid", "total", "balance_due",
    ]),
    "SaleItems": (SaleItem, [
        "id", "sale_id", "product_id", "product_name", "qty", "selling_price",
        "purchase_item_id", "line_total",
    ]),
    "StockMovements": (StockMovement, [
        "id", "date", "product_id", "product_name", "type", "qty",
        "reference_type", "reference_id", "note",
    ]),
    "Payments": (Payment, [
        "id", "sale_id", "invoice_no", "date", "amount", "payment_mode", "note",
    ]),
    "SaleReturns": (SaleReturn, [
        "id", "sale_id", "invoice_no", "applied_to_sale_id", "return_no", "date", "note", "refund_amount",
    ]),
    "SaleReturnItems": (SaleReturnItem, [
        "id", "sale_return_id", "sale_item_id", "product_name", "qty", "condition", "refund_amount",
    ]),
    "CustomerBrandDiscounts": (CustomerBrandDiscount, [
        "id", "customer_id", "customer_name", "brand_name", "discount_pct",
    ]),
    "MechanicBrandDiscounts": (MechanicBrandDiscount, [
        "id", "mechanic_id", "mechanic_name", "brand_name", "discount_pct",
    ]),
}

_DERIVED = {
    "supplier_name": lambda row: row.supplier.name if row.supplier else "",
    "customer_name": lambda row: row.customer.name if row.customer else "Walk-in",
    "mechanic_name": lambda row: row.mechanic.name if row.mechanic else "",
    "product_name": lambda row: row.product.product_name if row.product else "",
}


_MISSING = object()


def _cell_value(row, col):
    value = getattr(row, col, _MISSING)
    if value is not _MISSING:
        return value
    if col in _DERIVED:
        return _DERIVED[col](row)
    return ""


def _autofit(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, 10), 40)


def excel_path(app=None):
    app = app or current_app._get_current_object()
    return os.path.join(app.instance_path, EXCEL_FILENAME)


def _write_workbook(path):
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, (model, columns) in SHEETS.items():
        ws = wb.create_sheet(title=sheet_name)
        ws.append(columns)
        for row in model.query.order_by(model.id.asc()).all():
            ws.append([_cell_value(row, col) for col in columns])
        _autofit(ws)

    shop = ShopSettings.query.first()
    ws = wb.create_sheet(title="ShopSettings")
    ws.append(["shop_name", "address", "phone", "gstin"])
    if shop:
        ws.append([shop.shop_name, shop.address, shop.phone, shop.gstin])
    _autofit(ws)

    tmp_path = path + ".tmp"
    wb.save(tmp_path)
    os.replace(tmp_path, path)


def _db_path(app):
    uri = app.config["SQLALCHEMY_DATABASE_URI"]
    prefix = "sqlite:///"
    return uri[len(prefix):] if uri.startswith(prefix) else None


def _backup_to_cloud(app):
    """Best-effort copy of erp.db + erp_data.xlsx into CLOUD_BACKUP_DIR (see
    config.py) — typically a folder a cloud-sync client like Google Drive
    Desktop or OneDrive already watches, so the shop's data isn't only ever
    on one machine. Deliberately does NOT write directly into a live-synced
    folder from the app itself — a sync client can corrupt a SQLite file if
    it uploads mid-write. Instead this copies each file to a ".tmp" name in
    the backup folder first, then atomically renames it into place (same
    tmp-then-replace pattern _write_workbook uses), so the sync client only
    ever sees a complete file appear.

    Never raises and never flashes — this is a nice-to-have safety net, not
    something that should interrupt billing with a warning every time the
    shop is offline or the cloud folder isn't mounted, which is the normal
    case for an offline-first app."""
    backup_dir = app.config.get("CLOUD_BACKUP_DIR")
    if not backup_dir or not os.path.isdir(backup_dir):
        return
    try:
        sources = [p for p in (_db_path(app), excel_path(app)) if p and os.path.exists(p)]
        for src in sources:
            dst = os.path.join(backup_dir, os.path.basename(src))
            tmp_dst = dst + ".tmp"
            shutil.copy2(src, tmp_dst)
            os.replace(tmp_dst, dst)
    except Exception:
        current_app.logger.exception("Cloud backup failed")


def cloud_backup_status(app=None):
    """Snapshot for the dashboard banner and Settings > Cloud Backup page —
    {configured, connected, dir, last_synced}. 'connected' is the same check
    _backup_to_cloud() makes before copying (CLOUD_BACKUP_DIR set AND that
    folder currently exists) — it confirms the local sync folder is
    reachable, not that the cloud client has actually finished uploading.
    'last_synced' is read from the backed-up erp.db's own mtime rather than
    tracked separately, since that file is only ever written by a successful
    _backup_to_cloud() copy."""
    app = app or current_app._get_current_object()
    backup_dir = app.config.get("CLOUD_BACKUP_DIR")
    configured = bool(backup_dir)
    connected = configured and os.path.isdir(backup_dir)
    last_synced = None
    if connected:
        db_path = _db_path(app)
        if db_path:
            backed_up_db = os.path.join(backup_dir, os.path.basename(db_path))
            if os.path.exists(backed_up_db):
                last_synced = datetime.utcfromtimestamp(os.path.getmtime(backed_up_db))
    return {
        "configured": configured,
        "connected": connected,
        "dir": backup_dir,
        "last_synced": last_synced,
    }


def sync_to_excel():
    """Rewrite instance/erp_data.xlsx from the current DB state, then mirror
    both it and erp.db into CLOUD_BACKUP_DIR if configured. Never raises."""
    app = current_app._get_current_object()
    path = excel_path(app)
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        _write_workbook(path)
    except PermissionError:
        flash(
            "Saved to the database, but erp_data.xlsx is open elsewhere (e.g. Excel) "
            "so the Excel mirror couldn't be updated. It will catch up on the next change.",
            "warning",
        )
    except Exception:
        current_app.logger.exception("Excel sync failed")
        flash("Saved to the database, but updating the Excel mirror failed.", "warning")

    _backup_to_cloud(app)
