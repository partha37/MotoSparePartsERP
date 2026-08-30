import io
import re
from collections import defaultdict
from datetime import date, timedelta

from flask import Blueprint, render_template, request, send_file
from flask_login import login_required
from openpyxl import Workbook

from excel_sync import _autofit
from models import Sale, SaleItem, SaleReturn, Purchase, PurchaseItem, Product, Customer, Mechanic, Supplier, Brand

reports_bp = Blueprint("reports", __name__, url_prefix="/reports")


def _date_range_args():
    default_from = (date.today() - timedelta(days=30)).isoformat()
    default_to = date.today().isoformat()
    date_from = request.args.get("date_from", default_from)
    date_to = request.args.get("date_to", default_to)
    return date_from, date_to


def _safe_filename(name):
    """Strips characters Windows won't allow in a filename (the platform
    this app runs on) so a contact/product/brand/supplier name with a slash,
    colon, etc. in it can't break the download."""
    cleaned = re.sub(r'[\\/:*?"<>|]+', "_", name or "").strip()
    return cleaned or "export"


def _send_excel(sheets, filename):
    """sheets: list of (title, headers, rows) tuples — rows is an iterable of
    already Excel-safe values per row (date objects are fine, openpyxl writes
    them natively). One sheet per tuple, in order."""
    wb = Workbook()
    wb.remove(wb.active)
    for title, headers, rows in sheets:
        ws = wb.create_sheet(title=title[:31])
        ws.append(headers)
        for row in rows:
            ws.append(list(row))
        _autofit(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True, download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@reports_bp.route("/")
@login_required
def index():
    return render_template("reports/index.html")


def _outstanding_dues_rows():
    all_sales = Sale.query.order_by(Sale.date.asc(), Sale.id.asc()).all()
    return [s for s in all_sales if s.balance_due > 0]


@reports_bp.route("/outstanding-dues")
@login_required
def outstanding_dues():
    rows = _outstanding_dues_rows()
    grand_total = round(sum(s.balance_due for s in rows), 2)
    return render_template("reports/outstanding_dues.html", rows=rows, grand_total=grand_total)


@reports_bp.route("/outstanding-dues/export")
@login_required
def outstanding_dues_export():
    rows = _outstanding_dues_rows()
    sheet_rows = [
        [s.date, s.invoice_no, s.customer_display, s.mechanic.name if s.mechanic else "-",
         s.total, s.amount_paid, s.balance_due]
        for s in rows
    ]
    return _send_excel(
        [("Outstanding Dues", ["Date", "Invoice", "Customer", "Mechanic", "Total", "Paid", "Balance Due"], sheet_rows)],
        f"outstanding-dues-{date.today().isoformat()}.xlsx",
    )


def _compute_daily_sales(date_from, date_to, group_by):
    sales = (
        Sale.query.filter(Sale.date >= date_from, Sale.date <= date_to)
        .order_by(Sale.date.asc(), Sale.id.asc())
        .all()
    )

    by_period = defaultdict(lambda: {"count": 0, "total": 0.0})
    by_product = defaultdict(lambda: {"qty": 0, "revenue": 0.0, "cost": 0.0, "discount_amount": 0.0, "mrp_total": 0.0})
    total_revenue = 0.0
    total_cost = 0.0
    total_discount = 0.0
    total_mrp = 0.0

    for s in sales:
        period = _period_label(s.date, group_by)
        by_period[period]["count"] += 1
        by_period[period]["total"] += s.net_total

        for item in s.items:
            mrp = item.mrp_at_sale or 0
            unit_cost = (
                item.purchase_item.purchase_price if item.purchase_item
                else (item.product.actual_discounted_price or 0)
            )
            qty = item.net_qty
            line_revenue = item.net_line_total
            line_cost = round(qty * unit_cost, 2)
            line_discount = round((mrp - item.selling_price) * qty, 2) if mrp else 0
            product_key = f"{item.product.part_no} - {item.product.product_name}"

            by_product[product_key]["qty"] += qty
            by_product[product_key]["revenue"] += line_revenue
            by_product[product_key]["cost"] += line_cost
            by_product[product_key]["discount_amount"] += line_discount
            by_product[product_key]["mrp_total"] += mrp * qty

            total_revenue += line_revenue
            total_cost += line_cost
            total_discount += line_discount
            total_mrp += mrp * qty

    time_series = [(_period_display(p), v) for p, v in sorted(by_period.items(), key=lambda kv: kv[0])]

    product_rows = []
    for name, v in by_product.items():
        profit = round(v["revenue"] - v["cost"], 2)
        profit_pct = round(profit / v["revenue"] * 100, 2) if v["revenue"] else 0
        avg_discount_pct = round(v["discount_amount"] / v["mrp_total"] * 100, 2) if v["mrp_total"] else 0
        product_rows.append({
            "name": name, "qty": v["qty"], "revenue": round(v["revenue"], 2),
            "discount_amount": round(v["discount_amount"], 2), "avg_discount_pct": avg_discount_pct,
            "cost": round(v["cost"], 2), "profit": profit, "profit_pct": profit_pct,
        })
    product_rows.sort(key=lambda r: r["revenue"], reverse=True)

    total_profit = round(total_revenue - total_cost, 2)
    summary = {
        "invoices": len(sales),
        "total_revenue": round(total_revenue, 2),
        "total_discount": round(total_discount, 2),
        "avg_discount_pct": round(total_discount / total_mrp * 100, 2) if total_mrp else 0,
        "total_cost": round(total_cost, 2),
        "total_profit": total_profit,
        "profit_pct": round(total_profit / total_revenue * 100, 2) if total_revenue else 0,
    }
    return summary, time_series, product_rows


@reports_bp.route("/daily-sales")
@login_required
def daily_sales():
    date_from, date_to = _date_range_args()
    group_by = request.args.get("group", "day")
    if group_by not in ("day", "week"):
        group_by = "day"

    summary, time_series, product_rows = _compute_daily_sales(date_from, date_to, group_by)

    return render_template(
        "reports/daily_sales.html", summary=summary,
        time_series=time_series, product_rows=product_rows,
        date_from=date_from, date_to=date_to, group_by=group_by,
    )


@reports_bp.route("/daily-sales/export")
@login_required
def daily_sales_export():
    date_from, date_to = _date_range_args()
    group_by = request.args.get("group", "day")
    if group_by not in ("day", "week"):
        group_by = "day"

    summary, time_series, product_rows = _compute_daily_sales(date_from, date_to, group_by)
    period_header = "Date" if group_by == "day" else "Week"
    return _send_excel(
        [
            ("Sales Over Time", [period_header, "Invoices", "Total"],
             [[label, v["count"], v["total"]] for label, v in time_series]),
            ("By Product", ["Product", "Qty", "Revenue", "Discount Given", "Avg Discount %", "Cost", "Profit", "Profit %"],
             [[p["name"], p["qty"], p["revenue"], p["discount_amount"], p["avg_discount_pct"],
               p["cost"], p["profit"], p["profit_pct"]] for p in product_rows]),
        ],
        f"sales-overview-{date_from}_to_{date_to}.xlsx",
    )


def _compute_mechanic_wise(date_from, date_to):
    sales = Sale.query.filter(Sale.date >= date_from, Sale.date <= date_to).all()
    by_mechanic = defaultdict(lambda: {"id": None, "name": "No mechanic", "count": 0, "total": 0.0})
    for s in sales:
        key = s.mechanic_id or 0
        by_mechanic[key]["id"] = s.mechanic_id
        by_mechanic[key]["name"] = s.mechanic.name if s.mechanic else "No mechanic"
        by_mechanic[key]["count"] += 1
        by_mechanic[key]["total"] += s.net_total
    return sorted(by_mechanic.values(), key=lambda v: v["total"], reverse=True)


@reports_bp.route("/mechanic-wise")
@login_required
def mechanic_wise():
    date_from, date_to = _date_range_args()
    rows = _compute_mechanic_wise(date_from, date_to)
    return render_template(
        "reports/mechanic_wise.html", rows=rows, date_from=date_from, date_to=date_to
    )


@reports_bp.route("/mechanic-wise/export")
@login_required
def mechanic_wise_export():
    date_from, date_to = _date_range_args()
    rows = _compute_mechanic_wise(date_from, date_to)
    return _send_excel(
        [("Mechanic-wise Sales", ["Mechanic", "Invoices", "Total Spend"],
          [[r["name"], r["count"], r["total"]] for r in rows])],
        f"mechanic-wise-{date_from}_to_{date_to}.xlsx",
    )


def _compute_customer_wise(date_from, date_to):
    sales = Sale.query.filter(Sale.date >= date_from, Sale.date <= date_to).all()
    by_customer = defaultdict(lambda: {"id": None, "name": "Walk-in", "count": 0, "total": 0.0})
    for s in sales:
        key = s.customer_id or 0
        by_customer[key]["id"] = s.customer_id
        by_customer[key]["name"] = s.customer.name if s.customer else "Walk-in"
        by_customer[key]["count"] += 1
        by_customer[key]["total"] += s.net_total
    return sorted(by_customer.values(), key=lambda v: v["total"], reverse=True)


@reports_bp.route("/customer-wise")
@login_required
def customer_wise():
    date_from, date_to = _date_range_args()
    rows = _compute_customer_wise(date_from, date_to)
    return render_template(
        "reports/customer_wise.html", rows=rows, date_from=date_from, date_to=date_to
    )


@reports_bp.route("/customer-wise/export")
@login_required
def customer_wise_export():
    date_from, date_to = _date_range_args()
    rows = _compute_customer_wise(date_from, date_to)
    return _send_excel(
        [("Customer-wise Purchases", ["Customer", "Invoices", "Total Spend"],
          [[r["name"], r["count"], r["total"]] for r in rows])],
        f"customer-wise-{date_from}_to_{date_to}.xlsx",
    )


def _period_label(d, group_by):
    """Buckets a date into either its ISO date string (day) or ISO
    year-week string (week) — ISO sorts correctly as a plain string, which
    is why grouping/sorting always happens on this value. Never render this
    directly — pass it through _period_display() first, since "yyyy-mm-dd"
    doesn't match the dd-mm-yyyy format used everywhere else in the app."""
    if group_by == "week":
        iso_year, iso_week, _ = d.isocalendar()
        return f"{iso_year}-W{iso_week:02d}"
    return d.isoformat()


def _period_display(period):
    """Converts a _period_label() value to what should actually be shown to
    the user — dd-mm-yyyy for a day bucket (matching every other date in the
    app), left as-is for a week bucket ("2026-W29" isn't a single calendar
    date, so there's no dd-mm-yyyy equivalent to convert it to)."""
    try:
        return date.fromisoformat(period).strftime("%d-%m-%Y")
    except ValueError:
        return period


def _compute_contact_detail(kind, contact_id, date_from, date_to, group_by):
    if kind == "customer":
        contact = Customer.query.get_or_404(contact_id)
        sales_q = Sale.query.filter(Sale.customer_id == contact_id)
    else:
        contact = Mechanic.query.get_or_404(contact_id)
        sales_q = Sale.query.filter(Sale.mechanic_id == contact_id)

    sales = (
        sales_q.filter(Sale.date >= date_from, Sale.date <= date_to)
        .order_by(Sale.date.asc(), Sale.id.asc())
        .all()
    )

    by_period = defaultdict(lambda: {"count": 0, "total": 0.0})
    by_product = defaultdict(lambda: {"qty": 0, "revenue": 0.0, "cost": 0.0, "discount_amount": 0.0, "mrp_total": 0.0})
    item_rows = []
    total_revenue = 0.0
    total_cost = 0.0
    total_discount = 0.0
    total_mrp = 0.0

    for s in sales:
        period = _period_label(s.date, group_by)
        by_period[period]["count"] += 1
        by_period[period]["total"] += s.net_total

        for item in s.items:
            mrp = item.mrp_at_sale or 0
            unit_cost = (
                item.purchase_item.purchase_price if item.purchase_item
                else (item.product.actual_discounted_price or 0)
            )
            qty = item.net_qty
            line_revenue = item.net_line_total
            line_cost = round(qty * unit_cost, 2)
            line_discount = round((mrp - item.selling_price) * qty, 2) if mrp else 0
            product_key = f"{item.product.part_no} - {item.product.product_name}"

            item_rows.append({
                "date": s.date, "sale_id": s.id, "invoice_no": s.invoice_no,
                "product_name": product_key, "qty": item.qty, "returned_qty": item.returned_qty, "mrp": mrp,
                "discount_pct": item.discount_pct, "price": item.selling_price,
                "line_total": item.line_total,
            })

            by_product[product_key]["qty"] += qty
            by_product[product_key]["revenue"] += line_revenue
            by_product[product_key]["cost"] += line_cost
            by_product[product_key]["discount_amount"] += line_discount
            by_product[product_key]["mrp_total"] += mrp * qty

            total_revenue += line_revenue
            total_cost += line_cost
            total_discount += line_discount
            total_mrp += mrp * qty

    time_series = [(_period_display(p), v) for p, v in sorted(by_period.items(), key=lambda kv: kv[0])]
    item_rows.sort(key=lambda r: (r["date"], r["sale_id"]), reverse=True)

    product_rows = []
    for name, v in by_product.items():
        profit = round(v["revenue"] - v["cost"], 2)
        profit_pct = round(profit / v["revenue"] * 100, 2) if v["revenue"] else 0
        avg_discount_pct = round(v["discount_amount"] / v["mrp_total"] * 100, 2) if v["mrp_total"] else 0
        product_rows.append({
            "name": name, "qty": v["qty"], "revenue": round(v["revenue"], 2),
            "discount_amount": round(v["discount_amount"], 2), "avg_discount_pct": avg_discount_pct,
            "cost": round(v["cost"], 2), "profit": profit, "profit_pct": profit_pct,
        })
    product_rows.sort(key=lambda r: r["revenue"], reverse=True)

    total_profit = round(total_revenue - total_cost, 2)
    summary = {
        "invoices": len(sales),
        "total_revenue": round(total_revenue, 2),
        "total_discount": round(total_discount, 2),
        "avg_discount_pct": round(total_discount / total_mrp * 100, 2) if total_mrp else 0,
        "total_cost": round(total_cost, 2),
        "total_profit": total_profit,
        "profit_pct": round(total_profit / total_revenue * 100, 2) if total_revenue else 0,
    }

    return contact, summary, time_series, product_rows, item_rows


def _contact_detail(kind, contact_id):
    date_from, date_to = _date_range_args()
    group_by = request.args.get("group", "day")
    if group_by not in ("day", "week"):
        group_by = "day"

    contact, summary, time_series, product_rows, item_rows = _compute_contact_detail(
        kind, contact_id, date_from, date_to, group_by
    )

    return render_template(
        "reports/contact_detail.html",
        kind=kind, contact=contact, summary=summary,
        time_series=time_series, product_rows=product_rows, item_rows=item_rows,
        date_from=date_from, date_to=date_to, group_by=group_by,
    )


def _contact_detail_export(kind, contact_id):
    date_from, date_to = _date_range_args()
    group_by = request.args.get("group", "day")
    if group_by not in ("day", "week"):
        group_by = "day"

    contact, summary, time_series, product_rows, item_rows = _compute_contact_detail(
        kind, contact_id, date_from, date_to, group_by
    )
    period_header = "Date" if group_by == "day" else "Week"
    return _send_excel(
        [
            ("Over Time", [period_header, "Invoices", "Total"],
             [[label, v["count"], v["total"]] for label, v in time_series]),
            ("By Product", ["Product", "Qty", "Revenue", "Discount Given", "Avg Discount %", "Cost", "Profit", "Profit %"],
             [[p["name"], p["qty"], p["revenue"], p["discount_amount"], p["avg_discount_pct"],
               p["cost"], p["profit"], p["profit_pct"]] for p in product_rows]),
            ("Line Items", ["Date", "Invoice", "Product", "Qty", "Returned", "MRP", "Discount %", "Price", "Total"],
             [[r["date"], r["invoice_no"], r["product_name"], r["qty"], r["returned_qty"],
               r["mrp"], r["discount_pct"], r["price"], r["line_total"]] for r in item_rows]),
        ],
        f"{_safe_filename(contact.name)}-{date_from}_to_{date_to}.xlsx",
    )


@reports_bp.route("/customer/<int:contact_id>")
@login_required
def customer_detail(contact_id):
    return _contact_detail("customer", contact_id)


@reports_bp.route("/customer/<int:contact_id>/export")
@login_required
def customer_detail_export(contact_id):
    return _contact_detail_export("customer", contact_id)


@reports_bp.route("/mechanic/<int:contact_id>")
@login_required
def mechanic_detail(contact_id):
    return _contact_detail("mechanic", contact_id)


@reports_bp.route("/mechanic/<int:contact_id>/export")
@login_required
def mechanic_detail_export(contact_id):
    return _contact_detail_export("mechanic", contact_id)


def _compute_best_sellers(date_from, date_to):
    items = (
        SaleItem.query.join(Sale)
        .filter(Sale.date >= date_from, Sale.date <= date_to)
        .all()
    )

    by_product = defaultdict(lambda: {"id": None, "name": "", "qty": 0, "revenue": 0.0})
    for item in items:
        key = item.product_id
        by_product[key]["id"] = item.product_id
        by_product[key]["name"] = f"{item.product.part_no} - {item.product.product_name}"
        by_product[key]["qty"] += item.net_qty
        by_product[key]["revenue"] += item.net_line_total

    return sorted(by_product.values(), key=lambda v: v["qty"], reverse=True)


@reports_bp.route("/best-sellers")
@login_required
def best_sellers():
    date_from, date_to = _date_range_args()
    rows = _compute_best_sellers(date_from, date_to)
    return render_template(
        "reports/best_sellers.html", rows=rows, date_from=date_from, date_to=date_to
    )


@reports_bp.route("/best-sellers/export")
@login_required
def best_sellers_export():
    date_from, date_to = _date_range_args()
    rows = _compute_best_sellers(date_from, date_to)
    return _send_excel(
        [("Best Sellers", ["Product", "Qty Sold", "Revenue"],
          [[r["name"], r["qty"], r["revenue"]] for r in rows])],
        f"best-sellers-{date_from}_to_{date_to}.xlsx",
    )


def _compute_product_detail(product_id, date_from, date_to, group_by):
    product = Product.query.get_or_404(product_id)

    items = (
        SaleItem.query.join(Sale)
        .filter(SaleItem.product_id == product_id, Sale.date >= date_from, Sale.date <= date_to)
        .order_by(Sale.date.asc(), Sale.id.asc())
        .all()
    )

    by_period = defaultdict(lambda: {"count": 0, "qty": 0, "revenue": 0.0})
    by_contact = defaultdict(lambda: {"name": "", "qty": 0, "revenue": 0.0, "discount_amount": 0.0, "mrp_total": 0.0})
    item_rows = []
    total_revenue = 0.0
    total_cost = 0.0
    total_discount = 0.0
    total_mrp = 0.0
    total_qty = 0

    for item in items:
        s = item.sale
        period = _period_label(s.date, group_by)
        qty = item.net_qty
        line_revenue = item.net_line_total
        by_period[period]["count"] += 1
        by_period[period]["qty"] += qty
        by_period[period]["revenue"] += line_revenue

        mrp = item.mrp_at_sale or 0
        unit_cost = (
            item.purchase_item.purchase_price if item.purchase_item
            else (item.product.actual_discounted_price or 0)
        )
        line_cost = round(qty * unit_cost, 2)
        line_discount = round((mrp - item.selling_price) * qty, 2) if mrp else 0

        if s.mechanic_id:
            contact_key = ("mechanic", s.mechanic_id)
            contact_name = s.mechanic.name
        elif s.customer_id:
            contact_key = ("customer", s.customer_id)
            contact_name = s.customer.name
        else:
            contact_key = ("walkin", 0)
            contact_name = "Walk-in"

        item_rows.append({
            "date": s.date, "sale_id": s.id, "invoice_no": s.invoice_no,
            "contact_name": contact_name, "qty": item.qty, "returned_qty": item.returned_qty, "mrp": mrp,
            "discount_pct": item.discount_pct, "price": item.selling_price,
            "line_total": item.line_total,
        })

        by_contact[contact_key]["name"] = contact_name
        by_contact[contact_key]["qty"] += qty
        by_contact[contact_key]["revenue"] += line_revenue
        by_contact[contact_key]["discount_amount"] += line_discount
        by_contact[contact_key]["mrp_total"] += mrp * qty

        total_revenue += line_revenue
        total_cost += line_cost
        total_discount += line_discount
        total_mrp += mrp * qty
        total_qty += qty

    time_series = [(_period_display(p), v) for p, v in sorted(by_period.items(), key=lambda kv: kv[0])]
    item_rows.sort(key=lambda r: (r["date"], r["sale_id"]), reverse=True)

    contact_rows = []
    for v in by_contact.values():
        avg_discount_pct = round(v["discount_amount"] / v["mrp_total"] * 100, 2) if v["mrp_total"] else 0
        contact_rows.append({
            "name": v["name"], "qty": v["qty"], "revenue": round(v["revenue"], 2),
            "discount_amount": round(v["discount_amount"], 2), "avg_discount_pct": avg_discount_pct,
        })
    contact_rows.sort(key=lambda r: r["revenue"], reverse=True)

    total_profit = round(total_revenue - total_cost, 2)
    summary = {
        "qty_sold": total_qty,
        "total_revenue": round(total_revenue, 2),
        "total_discount": round(total_discount, 2),
        "avg_discount_pct": round(total_discount / total_mrp * 100, 2) if total_mrp else 0,
        "total_cost": round(total_cost, 2),
        "total_profit": total_profit,
        "profit_pct": round(total_profit / total_revenue * 100, 2) if total_revenue else 0,
        "current_stock": product.current_stock,
    }

    return product, summary, time_series, contact_rows, item_rows


def _product_detail(product_id):
    date_from, date_to = _date_range_args()
    group_by = request.args.get("group", "day")
    if group_by not in ("day", "week"):
        group_by = "day"

    product, summary, time_series, contact_rows, item_rows = _compute_product_detail(
        product_id, date_from, date_to, group_by
    )

    return render_template(
        "reports/product_detail.html",
        product=product, summary=summary,
        time_series=time_series, contact_rows=contact_rows, item_rows=item_rows,
        date_from=date_from, date_to=date_to, group_by=group_by,
    )


@reports_bp.route("/product/<int:product_id>")
@login_required
def product_detail(product_id):
    return _product_detail(product_id)


@reports_bp.route("/product/<int:product_id>/export")
@login_required
def product_detail_export(product_id):
    date_from, date_to = _date_range_args()
    group_by = request.args.get("group", "day")
    if group_by not in ("day", "week"):
        group_by = "day"

    product, summary, time_series, contact_rows, item_rows = _compute_product_detail(
        product_id, date_from, date_to, group_by
    )
    period_header = "Date" if group_by == "day" else "Week"
    return _send_excel(
        [
            ("Over Time", [period_header, "Invoices", "Qty", "Revenue"],
             [[label, v["count"], v["qty"], v["revenue"]] for label, v in time_series]),
            ("By Contact", ["Name", "Qty", "Revenue", "Discount Given", "Avg Discount %"],
             [[c["name"], c["qty"], c["revenue"], c["discount_amount"], c["avg_discount_pct"]] for c in contact_rows]),
            ("Line Items", ["Date", "Invoice", "Contact", "Qty", "Returned", "MRP", "Discount %", "Price", "Total"],
             [[r["date"], r["invoice_no"], r["contact_name"], r["qty"], r["returned_qty"],
               r["mrp"], r["discount_pct"], r["price"], r["line_total"]] for r in item_rows]),
        ],
        f"{_safe_filename(product.product_name)}-{date_from}_to_{date_to}.xlsx",
    )


def _compute_profit_margin(date_from, date_to):
    items = (
        SaleItem.query.join(Sale)
        .filter(Sale.date >= date_from, Sale.date <= date_to)
        .all()
    )

    by_product = defaultdict(lambda: {"qty": 0, "revenue": 0.0, "cost": 0.0})
    for item in items:
        key = f"{item.product.part_no} - {item.product.product_name}"
        # Use the actual batch's purchase price when known (accurate historical cost);
        # fall back to the product's current cost for sales made before batch tracking existed.
        unit_cost = (
            item.purchase_item.purchase_price if item.purchase_item
            else (item.product.actual_discounted_price or 0)
        )
        by_product[key]["qty"] += item.net_qty
        by_product[key]["revenue"] += item.net_line_total
        by_product[key]["cost"] += item.net_qty * unit_cost

    rows = []
    total_profit = 0.0
    for name, v in by_product.items():
        profit = round(v["revenue"] - v["cost"], 2)
        total_profit += profit
        rows.append((name, v["qty"], round(v["revenue"], 2), round(v["cost"], 2), profit))
    rows.sort(key=lambda r: r[4], reverse=True)
    return rows, round(total_profit, 2)


@reports_bp.route("/profit-margin")
@login_required
def profit_margin():
    date_from, date_to = _date_range_args()
    rows, total_profit = _compute_profit_margin(date_from, date_to)
    return render_template(
        "reports/profit_margin.html", rows=rows, total_profit=total_profit,
        date_from=date_from, date_to=date_to
    )


@reports_bp.route("/profit-margin/export")
@login_required
def profit_margin_export():
    date_from, date_to = _date_range_args()
    rows, total_profit = _compute_profit_margin(date_from, date_to)
    return _send_excel(
        [("Profit Margin", ["Product", "Qty Sold", "Revenue", "Cost", "Profit"], rows)],
        f"profit-margin-{date_from}_to_{date_to}.xlsx",
    )


def _compute_gst_report(date_from, date_to):
    """How much GST was actually paid on stock purchases — a per-product
    breakdown plus the raw purchase-line history, both scoped to the
    Purchase's business date (backdatable, same as every other report here).
    PurchaseItem.gst_amount is 0 for rows recorded before GST tracking
    existed (price_before_gst is null), so old purchases just don't
    contribute rather than skewing the totals."""
    items = (
        PurchaseItem.query.join(Purchase)
        .filter(Purchase.date >= date_from, Purchase.date <= date_to)
        .order_by(Purchase.date.desc(), PurchaseItem.id.desc())
        .all()
    )

    by_product = defaultdict(lambda: {"qty": 0, "pre_gst": 0.0, "gst": 0.0, "total": 0.0})
    item_rows = []
    total_pre_gst = 0.0
    total_gst = 0.0
    total_spend = 0.0

    for item in items:
        key = f"{item.product.part_no} - {item.product.product_name}" if item.product else "Unknown product"
        pre_gst_line = (item.price_before_gst or 0) * item.qty
        by_product[key]["qty"] += item.qty
        by_product[key]["pre_gst"] += pre_gst_line
        by_product[key]["gst"] += item.gst_amount
        by_product[key]["total"] += item.total
        total_pre_gst += pre_gst_line
        total_gst += item.gst_amount
        total_spend += item.total

        item_rows.append({
            "date": item.purchase.date,
            "purchase_id": item.purchase_id,
            "product_name": key,
            "qty": item.qty,
            "price_before_gst": item.price_before_gst,
            "gst_rate": item.gst_rate,
            "gst_amount": item.gst_amount,
            "purchase_price": item.purchase_price,
            "total": item.total,
        })

    product_rows = sorted(
        (
            {"name": name, "qty": v["qty"], "pre_gst": round(v["pre_gst"], 2),
             "gst": round(v["gst"], 2), "total": round(v["total"], 2)}
            for name, v in by_product.items()
        ),
        key=lambda r: r["gst"], reverse=True,
    )

    return product_rows, item_rows, round(total_pre_gst, 2), round(total_gst, 2), round(total_spend, 2)


@reports_bp.route("/gst")
@login_required
def gst_report():
    date_from, date_to = _date_range_args()
    product_rows, item_rows, total_pre_gst, total_gst, total_spend = _compute_gst_report(date_from, date_to)
    return render_template(
        "reports/gst_report.html",
        product_rows=product_rows, item_rows=item_rows,
        total_pre_gst=total_pre_gst, total_gst=total_gst,
        total_spend=total_spend, date_from=date_from, date_to=date_to,
    )


@reports_bp.route("/gst/export")
@login_required
def gst_report_export():
    date_from, date_to = _date_range_args()
    product_rows, item_rows, total_pre_gst, total_gst, total_spend = _compute_gst_report(date_from, date_to)
    return _send_excel(
        [
            ("By Product", ["Product", "Qty Purchased", "Price before GST", "GST Paid", "Total Spend"],
             [[p["name"], p["qty"], p["pre_gst"], p["gst"], p["total"]] for p in product_rows]),
            ("Purchase Line History", ["Date", "Product", "Qty", "Price before GST", "GST %", "GST Amount", "Price (incl. GST)", "Total"],
             [[r["date"], r["product_name"], r["qty"], r["price_before_gst"], r["gst_rate"],
               r["gst_amount"], r["purchase_price"], r["total"]] for r in item_rows]),
        ],
        f"gst-paid-{date_from}_to_{date_to}.xlsx",
    )


def _compute_supplier_summary(date_from, date_to):
    purchases = Purchase.query.filter(Purchase.date >= date_from, Purchase.date <= date_to).all()

    by_supplier = defaultdict(lambda: {"id": None, "name": "Unknown", "count": 0, "total": 0.0})
    for p in purchases:
        key = p.supplier_id or 0
        by_supplier[key]["id"] = p.supplier_id
        by_supplier[key]["name"] = p.supplier.name if p.supplier else "Unknown"
        by_supplier[key]["count"] += 1
        by_supplier[key]["total"] += p.total

    return sorted(by_supplier.values(), key=lambda v: v["total"], reverse=True)


@reports_bp.route("/supplier-summary")
@login_required
def supplier_summary():
    date_from, date_to = _date_range_args()
    rows = _compute_supplier_summary(date_from, date_to)
    return render_template(
        "reports/supplier_summary.html", rows=rows, date_from=date_from, date_to=date_to
    )


@reports_bp.route("/supplier-summary/export")
@login_required
def supplier_summary_export():
    date_from, date_to = _date_range_args()
    rows = _compute_supplier_summary(date_from, date_to)
    return _send_excel(
        [("Supplier Summary", ["Supplier", "Purchases", "Total Spend"],
          [[r["name"], r["count"], r["total"]] for r in rows])],
        f"supplier-summary-{date_from}_to_{date_to}.xlsx",
    )


def _compute_supplier_detail(supplier_id, date_from, date_to, group_by):
    supplier = Supplier.query.get_or_404(supplier_id)

    purchases = (
        Purchase.query.filter(
            Purchase.supplier_id == supplier_id, Purchase.date >= date_from, Purchase.date <= date_to
        )
        .order_by(Purchase.date.asc(), Purchase.id.asc())
        .all()
    )

    by_period = defaultdict(lambda: {"count": 0, "total": 0.0})
    by_product = defaultdict(lambda: {"qty": 0, "spend": 0.0, "discount_amount": 0.0, "mrp_total": 0.0})
    item_rows = []
    total_spend = 0.0
    total_discount = 0.0
    total_mrp = 0.0

    for p in purchases:
        period = _period_label(p.date, group_by)
        by_period[period]["count"] += 1
        by_period[period]["total"] += p.total

        for item in p.items:
            mrp = item.effective_mrp or 0
            line_spend = item.total
            line_discount = round((mrp - item.purchase_price) * item.qty, 2) if mrp else 0
            product_key = f"{item.product.part_no} - {item.product.product_name}"

            item_rows.append({
                "date": p.date, "purchase_id": p.id, "invoice_no": p.invoice_no,
                "product_name": product_key, "qty": item.qty, "mrp": mrp,
                "discount_pct": item.discount_pct, "price": item.purchase_price,
                "line_total": line_spend,
            })

            by_product[product_key]["qty"] += item.qty
            by_product[product_key]["spend"] += line_spend
            by_product[product_key]["discount_amount"] += line_discount
            by_product[product_key]["mrp_total"] += mrp * item.qty

            total_spend += line_spend
            total_discount += line_discount
            total_mrp += mrp * item.qty

    time_series = [(_period_display(p), v) for p, v in sorted(by_period.items(), key=lambda kv: kv[0])]
    item_rows.sort(key=lambda r: (r["date"], r["purchase_id"]), reverse=True)

    product_rows = []
    for name, v in by_product.items():
        avg_discount_pct = round(v["discount_amount"] / v["mrp_total"] * 100, 2) if v["mrp_total"] else 0
        product_rows.append({
            "name": name, "qty": v["qty"], "spend": round(v["spend"], 2),
            "discount_amount": round(v["discount_amount"], 2), "avg_discount_pct": avg_discount_pct,
        })
    product_rows.sort(key=lambda r: r["spend"], reverse=True)

    summary = {
        "purchases": len(purchases),
        "total_spend": round(total_spend, 2),
        "total_discount": round(total_discount, 2),
        "avg_discount_pct": round(total_discount / total_mrp * 100, 2) if total_mrp else 0,
    }

    return supplier, summary, time_series, product_rows, item_rows


@reports_bp.route("/supplier/<int:supplier_id>")
@login_required
def supplier_detail(supplier_id):
    date_from, date_to = _date_range_args()
    group_by = request.args.get("group", "day")
    if group_by not in ("day", "week"):
        group_by = "day"

    supplier, summary, time_series, product_rows, item_rows = _compute_supplier_detail(
        supplier_id, date_from, date_to, group_by
    )

    return render_template(
        "reports/supplier_detail.html",
        supplier=supplier, summary=summary,
        time_series=time_series, product_rows=product_rows, item_rows=item_rows,
        date_from=date_from, date_to=date_to, group_by=group_by,
    )


@reports_bp.route("/supplier/<int:supplier_id>/export")
@login_required
def supplier_detail_export(supplier_id):
    date_from, date_to = _date_range_args()
    group_by = request.args.get("group", "day")
    if group_by not in ("day", "week"):
        group_by = "day"

    supplier, summary, time_series, product_rows, item_rows = _compute_supplier_detail(
        supplier_id, date_from, date_to, group_by
    )
    period_header = "Date" if group_by == "day" else "Week"
    return _send_excel(
        [
            ("Over Time", [period_header, "Purchases", "Total"],
             [[label, v["count"], v["total"]] for label, v in time_series]),
            ("By Product", ["Product", "Qty", "Spend", "Discount Given", "Avg Discount %"],
             [[p["name"], p["qty"], p["spend"], p["discount_amount"], p["avg_discount_pct"]] for p in product_rows]),
            ("Line Items", ["Date", "Invoice", "Product", "Qty", "MRP", "Discount %", "Price", "Total"],
             [[r["date"], r["invoice_no"], r["product_name"], r["qty"], r["mrp"],
               r["discount_pct"], r["price"], r["line_total"]] for r in item_rows]),
        ],
        f"{_safe_filename(supplier.name)}-{date_from}_to_{date_to}.xlsx",
    )


def _compute_brand_wise(date_from, date_to):
    items = (
        SaleItem.query.join(Sale)
        .filter(Sale.date >= date_from, Sale.date <= date_to)
        .all()
    )

    by_brand = defaultdict(lambda: {"id": None, "name": "No brand", "qty": 0, "revenue": 0.0})
    for item in items:
        brand = item.product.brand
        key = brand.id if brand else 0
        by_brand[key]["id"] = brand.id if brand else None
        by_brand[key]["name"] = brand.name if brand else "No brand"
        by_brand[key]["qty"] += item.net_qty
        by_brand[key]["revenue"] += item.net_line_total

    return sorted(by_brand.values(), key=lambda v: v["revenue"], reverse=True)


@reports_bp.route("/brand-wise")
@login_required
def brand_wise():
    date_from, date_to = _date_range_args()
    rows = _compute_brand_wise(date_from, date_to)
    return render_template(
        "reports/brand_wise.html", rows=rows, date_from=date_from, date_to=date_to
    )


@reports_bp.route("/brand-wise/export")
@login_required
def brand_wise_export():
    date_from, date_to = _date_range_args()
    rows = _compute_brand_wise(date_from, date_to)
    return _send_excel(
        [("Brand-wise Sales", ["Brand", "Qty Sold", "Revenue"],
          [[r["name"], r["qty"], r["revenue"]] for r in rows])],
        f"brand-wise-{date_from}_to_{date_to}.xlsx",
    )


def _compute_brand_detail(brand_id, date_from, date_to, group_by):
    brand = Brand.query.get_or_404(brand_id)

    items = (
        SaleItem.query.join(Sale).join(Product)
        .filter(Product.brand_id == brand_id, Sale.date >= date_from, Sale.date <= date_to)
        .order_by(Sale.date.asc(), Sale.id.asc())
        .all()
    )

    by_period = defaultdict(lambda: {"count": 0, "revenue": 0.0})
    by_product = defaultdict(lambda: {"qty": 0, "revenue": 0.0, "cost": 0.0, "discount_amount": 0.0, "mrp_total": 0.0})
    item_rows = []
    total_revenue = 0.0
    total_cost = 0.0
    total_discount = 0.0
    total_mrp = 0.0
    total_qty = 0

    for item in items:
        s = item.sale
        period = _period_label(s.date, group_by)
        qty = item.net_qty
        line_revenue = item.net_line_total
        by_period[period]["count"] += 1
        by_period[period]["revenue"] += line_revenue

        mrp = item.mrp_at_sale or 0
        unit_cost = (
            item.purchase_item.purchase_price if item.purchase_item
            else (item.product.actual_discounted_price or 0)
        )
        line_cost = round(qty * unit_cost, 2)
        line_discount = round((mrp - item.selling_price) * qty, 2) if mrp else 0
        product_key = f"{item.product.part_no} - {item.product.product_name}"

        item_rows.append({
            "date": s.date, "sale_id": s.id, "invoice_no": s.invoice_no,
            "product_name": product_key, "qty": item.qty, "returned_qty": item.returned_qty, "mrp": mrp,
            "discount_pct": item.discount_pct, "price": item.selling_price,
            "line_total": item.line_total,
        })

        by_product[product_key]["qty"] += qty
        by_product[product_key]["revenue"] += line_revenue
        by_product[product_key]["cost"] += line_cost
        by_product[product_key]["discount_amount"] += line_discount
        by_product[product_key]["mrp_total"] += mrp * qty

        total_revenue += line_revenue
        total_cost += line_cost
        total_discount += line_discount
        total_mrp += mrp * qty
        total_qty += qty

    time_series = [(_period_display(p), v) for p, v in sorted(by_period.items(), key=lambda kv: kv[0])]
    item_rows.sort(key=lambda r: (r["date"], r["sale_id"]), reverse=True)

    product_rows = []
    for name, v in by_product.items():
        profit = round(v["revenue"] - v["cost"], 2)
        profit_pct = round(profit / v["revenue"] * 100, 2) if v["revenue"] else 0
        avg_discount_pct = round(v["discount_amount"] / v["mrp_total"] * 100, 2) if v["mrp_total"] else 0
        product_rows.append({
            "name": name, "qty": v["qty"], "revenue": round(v["revenue"], 2),
            "discount_amount": round(v["discount_amount"], 2), "avg_discount_pct": avg_discount_pct,
            "profit": profit, "profit_pct": profit_pct,
        })
    product_rows.sort(key=lambda r: r["revenue"], reverse=True)

    products = Product.query.filter(Product.brand_id == brand_id).order_by(Product.product_name.asc()).all()
    stock_value = round(sum((p.actual_discounted_price or 0) * (p.current_stock or 0) for p in products), 2)

    total_profit = round(total_revenue - total_cost, 2)
    summary = {
        "qty_sold": total_qty,
        "total_revenue": round(total_revenue, 2),
        "total_discount": round(total_discount, 2),
        "avg_discount_pct": round(total_discount / total_mrp * 100, 2) if total_mrp else 0,
        "total_cost": round(total_cost, 2),
        "total_profit": total_profit,
        "profit_pct": round(total_profit / total_revenue * 100, 2) if total_revenue else 0,
        "product_count": len(products),
        "stock_value": stock_value,
    }

    return brand, summary, products, time_series, product_rows, item_rows


@reports_bp.route("/brand/<int:brand_id>")
@login_required
def brand_detail(brand_id):
    date_from, date_to = _date_range_args()
    group_by = request.args.get("group", "day")
    if group_by not in ("day", "week"):
        group_by = "day"

    brand, summary, products, time_series, product_rows, item_rows = _compute_brand_detail(
        brand_id, date_from, date_to, group_by
    )

    return render_template(
        "reports/brand_detail.html",
        brand=brand, summary=summary, products=products,
        time_series=time_series, product_rows=product_rows, item_rows=item_rows,
        date_from=date_from, date_to=date_to, group_by=group_by,
    )


@reports_bp.route("/brand/<int:brand_id>/export")
@login_required
def brand_detail_export(brand_id):
    date_from, date_to = _date_range_args()
    group_by = request.args.get("group", "day")
    if group_by not in ("day", "week"):
        group_by = "day"

    brand, summary, products, time_series, product_rows, item_rows = _compute_brand_detail(
        brand_id, date_from, date_to, group_by
    )
    period_header = "Date" if group_by == "day" else "Week"
    return _send_excel(
        [
            ("Over Time", [period_header, "Invoices", "Revenue"],
             [[label, v["count"], v["revenue"]] for label, v in time_series]),
            ("By Product", ["Product", "Qty", "Revenue", "Discount Given", "Avg Discount %", "Profit", "Profit %"],
             [[p["name"], p["qty"], p["revenue"], p["discount_amount"], p["avg_discount_pct"],
               p["profit"], p["profit_pct"]] for p in product_rows]),
            ("Line Items", ["Date", "Invoice", "Product", "Qty", "Returned", "MRP", "Discount %", "Price", "Total"],
             [[r["date"], r["invoice_no"], r["product_name"], r["qty"], r["returned_qty"],
               r["mrp"], r["discount_pct"], r["price"], r["line_total"]] for r in item_rows]),
        ],
        f"{_safe_filename(brand.name)}-{date_from}_to_{date_to}.xlsx",
    )


def _compute_returns_report(date_from, date_to):
    """Returns & Defectives report — aggregates every SaleReturn/SaleReturnItem
    in range (by SaleReturn.date, the business return date) into a day-wise
    trend, a per-product resellable-vs-defective breakdown, and the raw list
    of return transactions. This is a different question from what the
    revenue/profit reports already answer by netting SaleItem.net_qty/
    net_line_total (they never show returned qty as its own number) — this
    report is where "what's coming back, and how much of it is defective"
    actually surfaces."""
    returns = (
        SaleReturn.query
        .filter(SaleReturn.date >= date_from, SaleReturn.date <= date_to)
        .order_by(SaleReturn.date.asc(), SaleReturn.id.asc())
        .all()
    )

    by_period = defaultdict(lambda: {"count": 0, "refund": 0.0})
    by_product = defaultdict(lambda: {"name": "", "resellable_qty": 0, "defective_qty": 0, "refund": 0.0})
    return_rows = []
    resellable_qty = 0
    resellable_value = 0.0
    defective_qty = 0
    defective_value = 0.0

    for r in returns:
        sale = r.sale
        if sale.mechanic_id:
            contact_name = sale.mechanic.name
        elif sale.customer_id:
            contact_name = sale.customer.name
        else:
            contact_name = "Walk-in"

        period = r.date.isoformat()
        by_period[period]["count"] += 1
        by_period[period]["refund"] += r.refund_amount

        row_resellable_qty = 0
        row_defective_qty = 0
        for item in r.items:
            key = item.product.id if item.product else 0
            by_product[key]["name"] = (
                f"{item.product.part_no} - {item.product.product_name}" if item.product else "Unknown product"
            )
            by_product[key]["refund"] += item.refund_amount
            if item.condition == "resellable":
                by_product[key]["resellable_qty"] += item.qty
                row_resellable_qty += item.qty
                resellable_qty += item.qty
                resellable_value += item.refund_amount
            else:
                by_product[key]["defective_qty"] += item.qty
                row_defective_qty += item.qty
                defective_qty += item.qty
                defective_value += item.refund_amount

        return_rows.append({
            "id": r.id, "sale_id": r.sale_id, "return_no": r.return_no,
            "date": r.date, "invoice_no": r.invoice_no, "contact_name": contact_name,
            "resellable_qty": row_resellable_qty, "defective_qty": row_defective_qty,
            "refund_amount": round(r.refund_amount, 2), "is_exchange": r.applied_to_sale_id is not None,
        })

    time_series = [(_period_display(p), v) for p, v in sorted(by_period.items(), key=lambda kv: kv[0])]
    return_rows.sort(key=lambda r: (r["date"], r["id"]), reverse=True)

    product_rows = sorted(
        (
            {
                "name": v["name"],
                "resellable_qty": v["resellable_qty"],
                "defective_qty": v["defective_qty"],
                "qty": v["resellable_qty"] + v["defective_qty"],
                "refund": round(v["refund"], 2),
            }
            for v in by_product.values()
        ),
        key=lambda r: r["qty"], reverse=True,
    )

    summary = {
        "return_count": len(returns),
        "total_refund": round(resellable_value + defective_value, 2),
        "resellable_qty": resellable_qty,
        "resellable_value": round(resellable_value, 2),
        "defective_qty": defective_qty,
        "defective_value": round(defective_value, 2),
    }

    return summary, time_series, product_rows, return_rows


@reports_bp.route("/returns")
@login_required
def returns_report():
    date_from, date_to = _date_range_args()
    summary, time_series, product_rows, return_rows = _compute_returns_report(date_from, date_to)
    return render_template(
        "reports/returns.html",
        summary=summary, time_series=time_series, product_rows=product_rows, return_rows=return_rows,
        date_from=date_from, date_to=date_to,
    )


@reports_bp.route("/returns/export")
@login_required
def returns_report_export():
    date_from, date_to = _date_range_args()
    summary, time_series, product_rows, return_rows = _compute_returns_report(date_from, date_to)
    return _send_excel(
        [
            ("Over Time", ["Date", "Returns", "Refund Value"],
             [[label, v["count"], v["refund"]] for label, v in time_series]),
            ("By Product", ["Product", "Resellable Qty", "Defective Qty", "Total Qty", "Refund Value"],
             [[p["name"], p["resellable_qty"], p["defective_qty"], p["qty"], p["refund"]] for p in product_rows]),
            ("Return History", ["Date", "Return No", "Invoice", "Customer / Mechanic", "Resellable Qty", "Defective Qty", "Refund", "Type"],
             [[r["date"], r["return_no"], r["invoice_no"], r["contact_name"], r["resellable_qty"], r["defective_qty"],
               r["refund_amount"], "Exchange" if r["is_exchange"] else "Return"] for r in return_rows]),
        ],
        f"returns-{date_from}_to_{date_to}.xlsx",
    )


def _attach_stock_status(products):
    """Flags each product 'out' / 'low' / 'ok' and whether it should be
    pre-checked by default on the Reorder report — the report lists the whole
    catalog (so any product can be picked for a purchase, not just ones
    currently low), but only out-of-stock/low-stock rows start pre-selected."""
    for p in products:
        if (p.current_stock or 0) <= 0:
            p.stock_status = "out"
        elif p.current_stock <= p.reorder_level:
            p.stock_status = "low"
        else:
            p.stock_status = "ok"
        p.needs_reorder = p.stock_status in ("out", "low")
    return products


@reports_bp.route("/low-stock")
@login_required
def low_stock():
    """The Reorder report — despite the route name (kept for URL stability),
    this lists the whole product catalog, not just low-stock items, so the
    shop owner can pick any product to reorder, not only what's flagged.
    Low/out-of-stock rows sort first and are pre-checked by default."""
    products = (
        Product.query
        .order_by(Product.current_stock.asc(), Product.product_name.asc())
        .all()
    )
    _attach_stock_status(products)
    return render_template("reports/low_stock.html", products=products)


@reports_bp.route("/low-stock/export", methods=["POST"])
@login_required
def low_stock_export():
    """Excel export for the Reorder report — exactly the checked products,
    whatever their stock status, or every low/out-of-stock product if
    nothing was checked (the page's own default selection)."""
    ids = [int(i) for i in request.form.getlist("product_ids[]") if i.isdigit()]
    if ids:
        query = Product.query.filter(Product.id.in_(ids))
    else:
        query = Product.query.filter(Product.current_stock <= Product.reorder_level)
    products = query.order_by(Product.current_stock.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Low Stock"
    ws.append(["Part No", "Product Name", "Brand", "Vehicle", "Current Stock", "Reorder Level", "Unit"])
    for p in products:
        ws.append([p.part_no, p.product_name, p.brand_name, p.vehicle_name or "", p.current_stock, p.reorder_level, p.unit])
    _autofit(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True,
        download_name=f"low-stock-{date.today().isoformat()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
